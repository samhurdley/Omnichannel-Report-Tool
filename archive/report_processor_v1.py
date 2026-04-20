import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import sys
import re

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', start_color='1F497D', end_color='1F497D')
SECTION_FILL = PatternFill('solid', start_color='2E4057', end_color='2E4057')
ALT_FILL     = PatternFill('solid', start_color='EEF2F7', end_color='EEF2F7')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
SECTION_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
DATA_FONT    = Font(name='Arial', size=10)
NUM   = '#,##0'
MONEY = '$#,##0.00'
PCT   = '0.00"%"'

def style_section_header(ws, row, title, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row, 1)
    cell.value = title
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 28

def style_col_headers(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 40

def style_data_row(ws, row, ncols):
    fill = ALT_FILL if row % 2 == 0 else None
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        if fill:
            cell.fill = fill
        cell.font = DATA_FONT
        cell.alignment = Alignment(
            horizontal='left' if c == 1 else 'center',
            vertical='center'
        )

def apply_formats(ws, row, fmts):
    for c, fmt in enumerate(fmts, 1):
        if fmt:
            ws.cell(row, c).number_format = fmt

# ── Helpers ───────────────────────────────────────────────────────────────────
def normalize(name):
    return re.sub(r'[\s_\-]+', '', str(name)).lower()

def find_client_config(filename, config_df):
    base = normalize(os.path.basename(filename))
    for _, row in config_df.iterrows():
        if normalize(row['Client Name']) in base:
            return row
    raise ValueError(
        f"No client in client_config.xlsx matched '{os.path.basename(filename)}'.\n"
        f"Clients available: {list(config_df['Client Name'])}"
    )

def find_col_by_key(csv_columns, config_row, key):
    val = config_row.get(key, '')
    if pd.notna(val) and str(val).strip():
        search = str(val).strip().lower()
        for c in csv_columns:
            if search in c.lower():
                return c
    return None

def find_conv_cols(csv_columns, config_row):
    cols = []
    for key in ['Conversion Column 1', 'Conversion Column 2']:
        col = find_col_by_key(csv_columns, config_row, key)
        if col and col not in cols:
            cols.append(col)
    return cols

def short_name(col):
    return re.sub(r'\s*\[.*$', '', col).strip()

CHANNEL_KEYWORDS = ['CTV', 'Video', 'Display', 'Audio']

def extract_channel(campaign_name):
    upper = str(campaign_name).upper()
    for ch in CHANNEL_KEYWORDS:
        if ch.upper() in upper:
            return ch
    return str(campaign_name)

def calc_metrics(imp, clk, spnd, conv, pcv):
    ctr  = clk  / imp  * 100  if imp  else 0
    cpc  = spnd / clk         if clk  else 0
    cpm  = spnd / imp  * 1000 if imp  else 0
    cvr  = conv / imp  * 100  if imp  else 0
    comp = pcv  / imp  * 100  if imp  else 0
    return ctr, cpc, cpm, cvr, comp

def roas(rev, spnd):
    return rev / spnd if spnd else 0

# ── Table spec builders ───────────────────────────────────────────────────────
def main_table_spec(label_col, conv_label, has_revenue):
    headers = [
        label_col, 'Impressions', 'Clicks', 'Spend',
        'CTR (%)', 'CPC', 'CPM',
        f'Conversions\n({conv_label})', 'Conv Rate\n(Conv/Impr %)',
        'Completion Rate\n(Video/CTV/Audio %)', 'Site Traffic',
    ]
    fmts = [None, NUM, NUM, MONEY, PCT, MONEY, MONEY, '#,##0.00', PCT, PCT, NUM]
    if has_revenue:
        headers += ['Revenue', 'ROAS']
        fmts    += [MONEY, '#,##0.00']
    return headers, fmts

def sites_table_spec(has_revenue):
    headers = ['Site', 'Impressions', 'Spend', 'CPM', 'Site Traffic']
    fmts    = [None, NUM, MONEY, MONEY, NUM]
    if has_revenue:
        headers += ['Revenue', 'ROAS']
        fmts    += [MONEY, '#,##0.00']
    return headers, fmts

# ── Section writers ───────────────────────────────────────────────────────────
def write_overall(ws, row, data, conv_label, has_revenue, ncols):
    imp, clk, spnd, conv, st, pcv, rev = data
    style_section_header(ws, row, 'OVERALL SUMMARY', ncols)
    row += 1

    ws.cell(row, 1).value = 'Metric'
    ws.cell(row, 2).value = 'Value'
    style_col_headers(ws, row, 2)
    row += 1

    _, cpc, cpm, cvr, _ = calc_metrics(imp, clk, spnd, conv, pcv)
    summary_rows = [
        ('Total Impressions',                      imp,  NUM),
        ('Total Clicks',                           clk,  NUM),
        ('Total Spend',                            spnd, MONEY),
        ('CPC',                                    cpc,  MONEY),
        ('CPM',                                    cpm,  MONEY),
        (f'Total Conversions ({conv_label})',      conv, '#,##0.00'),
        ('Conversion Rate (Conv / Impressions %)', cvr,  PCT),
        ('Site Traffic',                           st,   NUM),
    ]
    if has_revenue:
        summary_rows += [
            ('Total Revenue',  rev,            MONEY),
            ('ROAS',           roas(rev, spnd), '#,##0.00'),
        ]

    for label, val, fmt in summary_rows:
        ws.cell(row, 1).value = label
        ws.cell(row, 2).value = val
        ws.cell(row, 2).number_format = fmt
        style_data_row(ws, row, 2)
        row += 1

    return row + 2

def write_table_section(ws, row, title, headers, fmts, data_rows):
    ncols = len(headers)
    style_section_header(ws, row, title, ncols)
    row += 1

    for c, h in enumerate(headers, 1):
        ws.cell(row, c).value = h
    style_col_headers(ws, row, ncols)
    row += 1

    for data in data_rows:
        for c, val in enumerate(data, 1):
            ws.cell(row, c).value = val
        style_data_row(ws, row, ncols)
        apply_formats(ws, row, fmts)
        row += 1

    return row + 2

# ── Main ──────────────────────────────────────────────────────────────────────
def main(csv_path):
    csv_path = os.path.abspath(csv_path)
    folder = os.path.dirname(csv_path)

    config_path = os.path.join(folder, 'client_config.xlsx')
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"client_config.xlsx not found.\nExpected: {config_path}")

    config_df = pd.read_excel(config_path)
    config_row = find_client_config(csv_path, config_df)
    client_name = str(config_row['Client Name'])
    print(f"Matched client: {client_name}")

    df = pd.read_csv(csv_path)
    conv_cols = find_conv_cols(df.columns, config_row)
    if not conv_cols:
        raise ValueError(
            f"No conversion columns found for '{client_name}'.\n"
            "Check Conversion Column 1/2 in client_config.xlsx."
        )

    st_col  = find_col_by_key(df.columns, config_row, 'Site Traffic')
    rev_col = find_col_by_key(df.columns, config_row, 'Revenue')
    has_revenue = rev_col is not None

    conv_label = ' + '.join(short_name(c) for c in conv_cols)
    print(f"Conversion columns : {conv_cols}")
    print(f"Site Traffic column: {st_col or 'not found'}")
    print(f"Revenue column     : {rev_col or 'not configured'}")

    numeric_cols = (
        ['Impressions', 'Clicks', 'Player Completed Views', 'Advertiser Cost (Adv Currency)']
        + conv_cols
        + ([st_col]  if st_col  else [])
        + ([rev_col] if rev_col else [])
    )
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    df['_conv'] = df[conv_cols].sum(axis=1)
    df['_st']   = df[st_col]   if st_col  else 0
    df['_rev']  = df[rev_col]  if rev_col else 0
    df['_chan']  = df['Campaign'].apply(extract_channel)

    # Overall totals
    imp  = df['Impressions'].sum()
    clk  = df['Clicks'].sum()
    spnd = df['Advertiser Cost (Adv Currency)'].sum()
    conv = df['_conv'].sum()
    st   = df['_st'].sum()
    pcv  = df['Player Completed Views'].sum()
    rev  = df['_rev'].sum()

    def make_row(label, r):
        ctr, cpc, cpm, cvr, comp = calc_metrics(r.imp, r.clk, r.spnd, r.conv, r.pcv)
        row = [label, r.imp, r.clk, r.spnd, ctr, cpc, cpm, r.conv, cvr, comp, r.st]
        if has_revenue:
            row += [r.rev, roas(r.rev, r.spnd)]
        return row

    # By Channel
    grp_chan = df.groupby('_chan').agg(
        imp=('Impressions', 'sum'), clk=('Clicks', 'sum'),
        spnd=('Advertiser Cost (Adv Currency)', 'sum'),
        conv=('_conv', 'sum'), st=('_st', 'sum'),
        pcv=('Player Completed Views', 'sum'), rev=('_rev', 'sum')
    ).reset_index()
    camp_rows = [make_row(r['_chan'], r) for _, r in grp_chan.iterrows()]

    # All Creatives, sorted by conversions desc
    grp_cre = df.groupby('Creative').agg(
        imp=('Impressions', 'sum'), clk=('Clicks', 'sum'),
        spnd=('Advertiser Cost (Adv Currency)', 'sum'),
        conv=('_conv', 'sum'), st=('_st', 'sum'),
        pcv=('Player Completed Views', 'sum'), rev=('_rev', 'sum')
    ).reset_index().sort_values('conv', ascending=False)
    cre_rows = [make_row(r['Creative'], r) for _, r in grp_cre.iterrows()]

    # Top 10 Sites by site traffic (exclude [tail aggregate])
    df_sites = df[~df['Site'].str.contains(r'\[tail aggregate\]', case=False, na=False)]
    grp_site = df_sites.groupby('Site').agg(
        imp=('Impressions', 'sum'),
        spnd=('Advertiser Cost (Adv Currency)', 'sum'),
        st=('_st', 'sum'),
        rev=('_rev', 'sum')
    ).reset_index().sort_values('st', ascending=False).head(10)

    def make_site_row(r):
        cpm_ = r.spnd / r.imp * 1000 if r.imp else 0
        row = [r['Site'], r.imp, r.spnd, cpm_, r.st]
        if has_revenue:
            row += [r.rev, roas(r.rev, r.spnd)]
        return row

    site_rows = [make_site_row(r) for _, r in grp_site.iterrows()]

    # Build table specs
    main_hdrs, main_fmts = main_table_spec('Channel',  conv_label, has_revenue)
    cre_hdrs,  cre_fmts  = main_table_spec('Creative', conv_label, has_revenue)
    site_hdrs, site_fmts = sites_table_spec(has_revenue)
    ncols_main = len(main_hdrs)

    # Column widths (based on max ncols)
    base_widths = [58, 16, 14, 16, 14, 16, 16, 24, 18, 24, 16]
    if has_revenue:
        base_widths += [16, 14]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    for c, w in enumerate(base_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 1
    row = write_overall(ws, row, (imp, clk, spnd, conv, st, pcv, rev), conv_label, has_revenue, ncols_main)
    row = write_table_section(ws, row, 'BY CHANNEL',    main_hdrs, main_fmts, camp_rows)
    row = write_table_section(ws, row, 'ALL CREATIVES', cre_hdrs,  cre_fmts,  cre_rows)
    row = write_table_section(ws, row, 'TOP 10 SITES',  site_hdrs, site_fmts, site_rows)

    output_path = os.path.join(folder, f'{client_name}_summary.xlsx')
    wb.save(output_path)
    print(f"\nDone! Report saved to:\n{output_path}")

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage:  python process_report.py path/to/report.csv")
        sys.exit(1)
    try:
        main(sys.argv[1])
    except Exception as e:
        print(f"\nError: {e}")
        sys.exit(1)
