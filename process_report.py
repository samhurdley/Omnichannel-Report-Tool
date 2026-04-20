import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import sys
import re
import html as _html
import io
from datetime import datetime

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', start_color='1F497D', end_color='1F497D')
SECTION_FILL = PatternFill('solid', start_color='2E4057', end_color='2E4057')
ALT_FILL     = PatternFill('solid', start_color='EEF2F7', end_color='EEF2F7')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
SECTION_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
DATA_FONT    = Font(name='Arial', size=10)
NUM   = '#,##0'
MONEY = '$#,##0.00'
PCT   = '0.00"%"'

def style_section_header(ws, row, title, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row, 1)
    cell.value = title
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 28

def style_col_headers(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 40

def style_data_row(ws, row, ncols):
    fill = ALT_FILL if row % 2 == 0 else None
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        if fill:
            cell.fill = fill
        cell.font = DATA_FONT
        cell.alignment = Alignment(
            horizontal='left' if c == 1 else 'center',
            vertical='center'
        )

def apply_formats(ws, row, fmts):
    for c, fmt in enumerate(fmts, 1):
        if fmt:
            ws.cell(row, c).number_format = fmt

# ── Helpers ───────────────────────────────────────────────────────────────────
def normalize(name):
    return re.sub(r'[\s_\-]+', '', str(name)).lower()

def find_client_config(filename, config_df):
    base = normalize(os.path.basename(filename))
    for _, row in config_df.iterrows():
        if normalize(row['Client Name']) in base:
            return row
    raise ValueError(
        f"No client in client_config.xlsx matched '{os.path.basename(filename)}'.\n"
        f"Clients available: {list(config_df['Client Name'])}"
    )

def find_col_by_key(csv_columns, config_row, key):
    val = config_row.get(key, '')
    if pd.notna(val) and str(val).strip():
        search = str(val).strip().lower()
        for c in csv_columns:
            if search in c.lower():
                return c
    return None

def find_conv_cols(csv_columns, config_row):
    cols = []
    for key in ['Conversion Column 1', 'Conversion Column 2']:
        col = find_col_by_key(csv_columns, config_row, key)
        if col and col not in cols:
            cols.append(col)
    return cols

def short_name(col):
    return re.sub(r'\s*\[.*$', '', col).strip()

CHANNEL_KEYWORDS = ['CTV', 'Video', 'Display', 'Audio']

def extract_channel(campaign_name):
    upper = str(campaign_name).upper()
    for ch in CHANNEL_KEYWORDS:
        if ch.upper() in upper:
            return ch
    return str(campaign_name)

def calc_metrics(imp, clk, spnd, conv, pcv):
    ctr  = clk  / imp  * 100  if imp  else 0
    cpc  = spnd / clk         if clk  else 0
    cpm  = spnd / imp  * 1000 if imp  else 0
    cvr  = conv / imp  * 100  if imp  else 0
    comp = pcv  / imp  * 100  if imp  else 0
    return ctr, cpc, cpm, cvr, comp

def roas(rev, spnd):
    return rev / spnd if spnd else 0

# ── HTML helpers ──────────────────────────────────────────────────────────────
def extract_date_range(filename):
    dates = re.findall(r'\d{4}-\d{2}-\d{2}', os.path.basename(filename))
    if len(dates) >= 2:
        fmt = lambda d: datetime.strptime(d, '%Y-%m-%d').strftime('%-d %B %Y')
        return f"{fmt(dates[0])} – {fmt(dates[1])}"
    return ''

def extract_report_month(filename):
    dates = re.findall(r'\d{4}-\d{2}-\d{2}', os.path.basename(filename))
    if dates:
        return datetime.strptime(dates[0], '%Y-%m-%d').strftime('%B %Y')
    return ''

def _h(s):   return _html.escape(str(s))
def _n(v):   return f"{v:,.0f}"
def _m(v):   return f"${v:,.2f}"
def _p(v):   return f"{v:.2f}%"
def _rx(v):  return f"{v:.2f}x"

_KPI_CYCLE = ['#1F497D', '#2E86AB', '#48A999', '#E8703A',
              '#7B5EA7', '#F4845F', '#3BB273', '#2E4057']

def _kpi_card(label, value, color='#1F497D'):
    return f'''
        <div class="kpi-card" style="border-left:4px solid {color}">
          <div class="kpi-value">{value}</div>
          <div class="kpi-label">{label}</div>
        </div>'''

def _th(cells, right_from=1):
    ths = ''
    for i, c in enumerate(cells):
        cls = ' class="right"' if i >= right_from else ''
        ths += f'<th{cls}>{_h(c)}</th>'
    return f'<tr>{ths}</tr>'

def _td_row(cells, right_from=1, highlight=False):
    cls = ' class="highlight"' if highlight else ''
    tds = ''
    for i, c in enumerate(cells):
        align = ' class="right"' if i >= right_from else ''
        # values that are already safe HTML (e.g. the NA span) skip escaping
        safe = c if (isinstance(c, str) and c.startswith('<')) else _h(c)
        tds += f'<td{align}>{safe}</td>'
    return f'<tr{cls}>{tds}</tr>'

def generate_html(csv_path, client_name, conv_label, has_revenue,
                  totals, grp_chan, grp_cre, grp_site):
    imp, clk, spnd, conv, st, pcv, rev = totals
    _, cpc, cpm, cvr, _ = calc_metrics(imp, clk, spnd, conv, pcv)
    ctr_overall = clk / imp * 100 if imp else 0
    report_month = extract_report_month(csv_path)
    title = f"{client_name} Omnichannel Report — {report_month}"

    NA = '<span style="color:#aab0c4;font-size:11px">N/A</span>'

    # ── KPI cards ─────────────────────────────────────────────────────────────
    kpi_defs = [
        ('Total Impressions',  _n(imp)),
        ('Total Clicks',       _n(clk)),
        ('Total Spend',        _m(spnd)),
        ('CTR',                _p(ctr_overall)),
        ('CPC',                _m(cpc)),
        ('CPM',                _m(cpm)),
        (f'Conversions ({conv_label})', _n(conv)),
        ('Conversion Rate',    _p(cvr)),
    ]
    if has_revenue:
        kpi_defs += [
            ('Attributed Revenue', _m(rev)),
            ('ROAS',               _rx(roas(rev, spnd))),
        ]
    kpi_cards = ''.join(
        _kpi_card(lbl, val, _KPI_CYCLE[i % len(_KPI_CYCLE)])
        for i, (lbl, val) in enumerate(kpi_defs)
    )
    kpi_cols = 5 if has_revenue else 4

    # ── By Channel table ──────────────────────────────────────────────────────
    NO_CLICKS = {'CTV', 'Audio'}
    NO_COMP   = {'Display'}

    chan_hdrs = ['Channel', 'Impressions', 'Clicks', 'Spend',
                 'CTR', 'CPC', 'CPM', 'Conversions', 'Conv Rate',
                 'Completion Rate', 'Attributed Site Traffic']
    if has_revenue:
        chan_hdrs += ['Attributed Revenue', 'ROAS']

    chan_head = _th(chan_hdrs)
    chan_body = ''
    for _, r in grp_chan.iterrows():
        ctr_, cpc_, cpm_, cvr_, comp_ = calc_metrics(r.imp, r.clk, r.spnd, r.conv, r.pcv)
        ch = str(r['_chan'])
        cells = [
            ch,
            _n(r.imp),
            NA if ch in NO_CLICKS else _n(r.clk),
            _m(r.spnd),
            NA if ch in NO_CLICKS else _p(ctr_),
            NA if ch in NO_CLICKS else _m(cpc_),
            _m(cpm_),
            _n(r.conv),
            _p(cvr_),
            NA if ch in NO_COMP else _p(comp_),
            _n(r.st),
        ]
        if has_revenue:
            cells += [_m(r.rev), _rx(roas(r.rev, r.spnd))]
        chan_body += _td_row(cells)

    # ── Top 10 Creatives by attributed site traffic ───────────────────────────
    top10_cre = grp_cre.nlargest(10, 'st')

    cre_hdrs = ['Creative', 'Impressions', 'Spend', 'CTR', 'CPC',
                'Conversions', 'Conv Rate', 'Completion Rate',
                'Attributed Site Traffic']
    if has_revenue:
        cre_hdrs += ['Attributed Revenue', 'ROAS']

    cre_head = _th(cre_hdrs)
    cre_body = ''
    for i, (_, r) in enumerate(top10_cre.iterrows()):
        ctr_, cpc_, cpm_, cvr_, comp_ = calc_metrics(r.imp, r.clk, r.spnd, r.conv, r.pcv)
        comp_v = NA if comp_ == 0 else _p(comp_)
        cells = [r['Creative'], _n(r.imp), _m(r.spnd), _p(ctr_), _m(cpc_),
                 _n(r.conv), _p(cvr_), comp_v, _n(r.st)]
        if has_revenue:
            cells += [_m(r.rev), _rx(roas(r.rev, r.spnd))]
        cre_body += _td_row(cells, highlight=(i == 0))

    # ── Top 10 Sites ──────────────────────────────────────────────────────────
    site_hdrs = ['Site', 'Impressions', 'Spend', 'CPM', 'Attributed Site Traffic']
    if has_revenue:
        site_hdrs += ['Attributed Revenue', 'ROAS']

    site_head = _th(site_hdrs)
    site_body = ''
    for _, r in grp_site.iterrows():
        cpm_ = r.spnd / r.imp * 1000 if r.imp else 0
        cells = [r['Site'], _n(r.imp), _m(r.spnd), _m(cpm_), _n(r.st)]
        if has_revenue:
            cells += [_m(r.rev), _rx(roas(r.rev, r.spnd))]
        site_body += _td_row(cells)

    return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{_h(title)}</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
    background: #f0f2f6;
    color: #1a1e2e;
    font-size: 14px;
    line-height: 1.5;
  }}

  /* ── Header ── */
  .report-header {{
    background: #1F497D;
    color: #fff;
    padding: 36px 48px 32px;
  }}
  .report-header h1 {{
    font-size: 26px;
    font-weight: 700;
    letter-spacing: -0.3px;
  }}
  .report-header .subtitle {{
    font-size: 13px;
    opacity: 0.75;
    margin-top: 6px;
    letter-spacing: 0.02em;
  }}

  /* ── Layout ── */
  .content {{
    max-width: 1280px;
    margin: 0 auto;
    padding: 32px 32px 64px;
  }}

  .section-label {{
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #6b7a99;
    margin: 40px 0 14px;
  }}
  .section-label:first-child {{ margin-top: 0; }}

  /* ── KPI cards ── */
  .kpi-grid {{
    display: grid;
    grid-template-columns: repeat({kpi_cols}, 1fr);
    gap: 14px;
  }}
  .kpi-card {{
    background: #fff;
    border-radius: 10px;
    padding: 22px 20px 18px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.07), 0 4px 12px rgba(0,0,0,0.04);
  }}
  .kpi-value {{
    font-size: 26px;
    font-weight: 700;
    color: #1a1e2e;
    font-variant-numeric: tabular-nums;
    line-height: 1.1;
  }}
  .kpi-label {{
    font-size: 11px;
    color: #6b7a99;
    margin-top: 7px;
    text-transform: uppercase;
    letter-spacing: 0.07em;
    font-weight: 600;
  }}

  /* ── Tables ── */
  .table-wrap {{
    background: #fff;
    border-radius: 10px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.07), 0 4px 12px rgba(0,0,0,0.04);
    overflow-x: auto;   /* horizontal scroll if viewport is narrow */
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 12.5px;
    min-width: 600px;   /* prevents columns collapsing below readable width */
  }}
  thead tr {{ background: #1F497D; }}
  th {{
    color: #fff;
    padding: 11px 14px;
    font-weight: 600;
    font-size: 10.5px;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    white-space: normal;
    min-width: 80px;
    vertical-align: bottom;
    line-height: 1.3;
    text-align: right;        /* all headers right by default */
  }}
  th:first-child {{
    min-width: 160px;
    text-align: left;         /* label column always left */
  }}
  td {{
    padding: 10px 14px;
    border-bottom: 1px solid #eef0f5;
    color: #2c3350;
    vertical-align: middle;
    text-align: left;         /* explicit default */
  }}
  td:first-child {{
    word-break: break-word;
    overflow-wrap: anywhere;
    max-width: 260px;
    line-height: 1.4;
    text-align: left;
  }}
  td.right {{
    text-align: right;
    white-space: nowrap;
    font-variant-numeric: tabular-nums;
  }}
  tr:last-child td {{ border-bottom: none; }}
  tbody tr:nth-child(even) td {{ background: #f7f8fc; }}
  tbody tr.highlight td {{ background: #fffbeb; font-weight: 600; }}
  tbody tr:hover td {{ background: #eef3fb; }}

  /* ── Conversions note ── */
  .conv-note {{
    display: inline-block;
    margin-top: 14px;
    padding: 8px 14px;
    background: #fff;
    border-left: 3px solid #1F497D;
    border-radius: 0 6px 6px 0;
    font-size: 12px;
    color: #6b7a99;
    box-shadow: 0 1px 3px rgba(0,0,0,0.06);
  }}
  .conv-note strong {{
    color: #1a1e2e;
    font-weight: 600;
  }}

  /* ── Footer ── */
  .footer {{
    text-align: center;
    font-size: 11px;
    color: #9aa3bc;
    margin-top: 48px;
    padding-top: 20px;
    border-top: 1px solid #e4e7f0;
  }}

  /* ── Print ── */
  @media print {{
    body {{ background: #fff; }}
    .report-header {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .kpi-card {{
      box-shadow: none; border: 1px solid #dde2ef;
      -webkit-print-color-adjust: exact; print-color-adjust: exact;
    }}
    .table-wrap {{
      box-shadow: none; border: 1px solid #dde2ef;
      overflow-x: visible;
      page-break-inside: avoid;
    }}
    thead tr {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .kpi-grid {{ grid-template-columns: repeat({kpi_cols}, 1fr); }}
    table {{ min-width: unset; font-size: 9px; }}
    th, td {{ padding: 6px 8px; }}
    td:first-child {{ max-width: 180px; }}
  }}
</style>
</head>
<body>

<div class="report-header">
  <h1>{_h(title)}</h1>
  <div class="subtitle">{_h(extract_date_range(csv_path))}</div>
</div>

<div class="content">

  <div class="section-label">Key Performance Indicators</div>
  <div class="kpi-grid">
    {kpi_cards}
  </div>
  <div class="conv-note">Conversions tracked: <strong>{_h(conv_label)}</strong></div>

  <div class="section-label">Performance by Channel</div>
  <div class="table-wrap">
    <table>
      <thead>{chan_head}</thead>
      <tbody>{chan_body}</tbody>
    </table>
  </div>

  <div class="section-label">Top 10 Creatives by Attributed Site Traffic</div>
  <div class="table-wrap">
    <table>
      <thead>{cre_head}</thead>
      <tbody>{cre_body}</tbody>
    </table>
  </div>

  <div class="section-label">Top 10 Sites by Attributed Site Traffic</div>
  <div class="table-wrap">
    <table>
      <thead>{site_head}</thead>
      <tbody>{site_body}</tbody>
    </table>
  </div>

  <div class="footer">Confidential — prepared for {_h(client_name)}</div>

</div>
</body>
</html>'''

# ── Table spec builders ───────────────────────────────────────────────────────
def main_table_spec(label_col, conv_label, has_revenue):
    headers = [
        label_col, 'Impressions', 'Clicks', 'Spend',
        'CTR (%)', 'CPC', 'CPM',
        f'Conversions\n({conv_label})', 'Conv Rate\n(Conv/Impr %)',
        'Completion Rate\n(Video/CTV/Audio %)', 'Site Traffic',
    ]
    fmts = [None, NUM, NUM, MONEY, PCT, MONEY, MONEY, '#,##0.00', PCT, PCT, NUM]
    if has_revenue:
        headers += ['Revenue', 'ROAS']
        fmts    += [MONEY, '#,##0.00']
    return headers, fmts

def sites_table_spec(has_revenue):
    headers = ['Site', 'Impressions', 'Spend', 'CPM', 'Site Traffic']
    fmts    = [None, NUM, MONEY, MONEY, NUM]
    if has_revenue:
        headers += ['Revenue', 'ROAS']
        fmts    += [MONEY, '#,##0.00']
    return headers, fmts

# ── Section writers ───────────────────────────────────────────────────────────
def write_overall(ws, row, data, conv_label, has_revenue, ncols):
    imp, clk, spnd, conv, st, pcv, rev = data
    style_section_header(ws, row, 'OVERALL SUMMARY', ncols)
    row += 1

    ws.cell(row, 1).value = 'Metric'
    ws.cell(row, 2).value = 'Value'
    style_col_headers(ws, row, 2)
    row += 1

    _, cpc, cpm, cvr, _ = calc_metrics(imp, clk, spnd, conv, pcv)
    summary_rows = [
        ('Total Impressions',                      imp,  NUM),
        ('Total Clicks',                           clk,  NUM),
        ('Total Spend',                            spnd, MONEY),
        ('CPC',                                    cpc,  MONEY),
        ('CPM',                                    cpm,  MONEY),
        (f'Total Conversions ({conv_label})',      conv, '#,##0.00'),
        ('Conversion Rate (Conv / Impressions %)', cvr,  PCT),
        ('Site Traffic',                           st,   NUM),
    ]
    if has_revenue:
        summary_rows += [
            ('Total Revenue',  rev,            MONEY),
            ('ROAS',           roas(rev, spnd), '#,##0.00'),
        ]

    for label, val, fmt in summary_rows:
        ws.cell(row, 1).value = label
        ws.cell(row, 2).value = val
        ws.cell(row, 2).number_format = fmt
        style_data_row(ws, row, 2)
        row += 1

    return row + 2

def write_table_section(ws, row, title, headers, fmts, data_rows):
    ncols = len(headers)
    style_section_header(ws, row, title, ncols)
    row += 1

    for c, h in enumerate(headers, 1):
        ws.cell(row, c).value = h
    style_col_headers(ws, row, ncols)
    row += 1

    for data in data_rows:
        for c, val in enumerate(data, 1):
            ws.cell(row, c).value = val
        style_data_row(ws, row, ncols)
        apply_formats(ws, row, fmts)
        row += 1

    return row + 2

# ── Core processor (used by both CLI and Streamlit) ───────────────────────────
def process_csv(csv_bytes, csv_filename, config_df):
    """
    Process one CSV report.
    Returns (client_name, excel_bytes, html_str).
    Raises ValueError/FileNotFoundError on bad input.
    """
    config_row  = find_client_config(csv_filename, config_df)
    client_name = str(config_row['Client Name'])

    df = pd.read_csv(io.BytesIO(csv_bytes))
    conv_cols = find_conv_cols(df.columns, config_row)
    if not conv_cols:
        raise ValueError(
            f"No conversion columns found for '{client_name}'. "
            "Check Conversion Column 1/2 in client_config.xlsx."
        )

    st_col      = find_col_by_key(df.columns, config_row, 'Site Traffic')
    rev_col     = find_col_by_key(df.columns, config_row, 'Revenue')
    has_revenue = rev_col is not None
    conv_label  = ' + '.join(short_name(c) for c in conv_cols)

    numeric_cols = (
        ['Impressions', 'Clicks', 'Player Completed Views', 'Advertiser Cost (Adv Currency)']
        + conv_cols
        + ([st_col]  if st_col  else [])
        + ([rev_col] if rev_col else [])
    )
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    df['_conv'] = df[conv_cols].sum(axis=1)
    df['_st']   = df[st_col]  if st_col  else 0
    df['_rev']  = df[rev_col] if rev_col else 0
    df['_chan']  = df['Campaign'].apply(extract_channel)

    imp  = df['Impressions'].sum()
    clk  = df['Clicks'].sum()
    spnd = df['Advertiser Cost (Adv Currency)'].sum()
    conv = df['_conv'].sum()
    st   = df['_st'].sum()
    pcv  = df['Player Completed Views'].sum()
    rev  = df['_rev'].sum()

    def make_row(label, r):
        ctr, cpc, cpm, cvr, comp = calc_metrics(r.imp, r.clk, r.spnd, r.conv, r.pcv)
        row = [label, r.imp, r.clk, r.spnd, ctr, cpc, cpm, r.conv, cvr, comp, r.st]
        if has_revenue:
            row += [r.rev, roas(r.rev, r.spnd)]
        return row

    grp_chan = df.groupby('_chan').agg(
        imp=('Impressions', 'sum'), clk=('Clicks', 'sum'),
        spnd=('Advertiser Cost (Adv Currency)', 'sum'),
        conv=('_conv', 'sum'), st=('_st', 'sum'),
        pcv=('Player Completed Views', 'sum'), rev=('_rev', 'sum')
    ).reset_index()
    camp_rows = [make_row(r['_chan'], r) for _, r in grp_chan.iterrows()]

    grp_cre = df.groupby('Creative').agg(
        imp=('Impressions', 'sum'), clk=('Clicks', 'sum'),
        spnd=('Advertiser Cost (Adv Currency)', 'sum'),
        conv=('_conv', 'sum'), st=('_st', 'sum'),
        pcv=('Player Completed Views', 'sum'), rev=('_rev', 'sum')
    ).reset_index().sort_values('conv', ascending=False)
    cre_rows = [make_row(r['Creative'], r) for _, r in grp_cre.iterrows()]

    df_sites = df[~df['Site'].str.contains(r'\[tail aggregate\]', case=False, na=False)]
    grp_site = df_sites.groupby('Site').agg(
        imp=('Impressions', 'sum'),
        spnd=('Advertiser Cost (Adv Currency)', 'sum'),
        st=('_st', 'sum'),
        rev=('_rev', 'sum')
    ).reset_index().sort_values('st', ascending=False).head(10)

    def make_site_row(r):
        cpm_ = r.spnd / r.imp * 1000 if r.imp else 0
        row = [r['Site'], r.imp, r.spnd, cpm_, r.st]
        if has_revenue:
            row += [r.rev, roas(r.rev, r.spnd)]
        return row

    site_rows = [make_site_row(r) for _, r in grp_site.iterrows()]

    main_hdrs, main_fmts = main_table_spec('Channel',  conv_label, has_revenue)
    cre_hdrs,  cre_fmts  = main_table_spec('Creative', conv_label, has_revenue)
    site_hdrs, site_fmts = sites_table_spec(has_revenue)
    ncols_main = len(main_hdrs)

    base_widths = [58, 16, 14, 16, 14, 16, 16, 24, 18, 24, 16]
    if has_revenue:
        base_widths += [16, 14]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'
    for c, w in enumerate(base_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    row = 1
    row = write_overall(ws, row, (imp, clk, spnd, conv, st, pcv, rev), conv_label, has_revenue, ncols_main)
    row = write_table_section(ws, row, 'BY CHANNEL',    main_hdrs, main_fmts, camp_rows)
    row = write_table_section(ws, row, 'ALL CREATIVES', cre_hdrs,  cre_fmts,  cre_rows)
    row = write_table_section(ws, row, 'TOP 10 SITES',  site_hdrs, site_fmts, site_rows)

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()

    html_str = generate_html(
        csv_filename, client_name, conv_label, has_revenue,
        (imp, clk, spnd, conv, st, pcv, rev),
        grp_chan, grp_cre, grp_site
    )

    return client_name, excel_bytes, html_str


# ── CLI entry point ────────────────────────────────────────────────────────────
def main(csv_path):
    csv_path = os.path.abspath(csv_path)
    folder   = os.path.dirname(csv_path)

    config_path = os.path.join(folder, 'client_config.xlsx')
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"client_config.xlsx not found.\nExpected: {config_path}")

    config_df = pd.read_excel(config_path)
    with open(csv_path, 'rb') as f:
        csv_bytes = f.read()

    client_name, excel_bytes, html_str = process_csv(
        csv_bytes, os.path.basename(csv_path), config_df
    )
    print(f"Matched client: {client_name}")

    xlsx_path = os.path.join(folder, f'{client_name}_summary.xlsx')
    with open(xlsx_path, 'wb') as f:
        f.write(excel_bytes)
    print(f"Excel saved : {xlsx_path}")

    html_path = os.path.join(folder, f'{client_name}_report.html')
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_str)
    print(f"HTML saved  : {html_path}")


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage:  python process_report.py path/to/report.csv")
        sys.exit(1)
    try:
        main(sys.argv[1])
    except Exception as e:
        print(f"\nError: {e}")
        sys.exit(1)
